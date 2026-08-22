"""xlsx_read / xlsx_append / xlsx_write edit a workbook in place and obey the group boundary.

Run:  python tests/xlsx_actions_test.py     (exit 0 = pass)

Why this file exists. The only pre-existing way to write an .xlsx was
FileService.write_sheet(), which rebuilds the whole workbook from a cell grid:
every sheet that was not sent back disappears, and so do styles, column widths
and formulas. An agent that "adds one row to the upload template" through that
path would quietly wreck the template. The new methods open the workbook with
openpyxl and touch only the cells they were asked to, so this test asserts on
what SURVIVES the edit (other tab, bold header, border, width, formula), not
just on what was written.

The second half drives the extension handlers with a stand-in
tubecli.core.group_context that follows the spec signatures (effective_groups,
resolve_xlsx, allows, load). With groups, a path outside the manifest must be
refused even when the sandbox would allow it, a path inside the manifest must
work even when the sandbox would refuse it, and the entry's access must gate
append vs write. Without groups the classic sandbox applies.
"""
import asyncio
import os
import shutil
import sys
import types

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except AttributeError:
    pass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Side

from tubecli.extensions.file_manager.file_service import FileService, file_service

PASS = FAIL = 0


def check(name, ok, detail=""):
    global PASS, FAIL
    if ok:
        PASS += 1
        print(f"[PASS] {name}" + (f"  ({detail})" if detail else ""))
    else:
        FAIL += 1
        print(f"[FAIL] {name}" + (f"  ({detail})" if detail else ""))


def expect_raises(name, exc, fn, *args, **kw):
    try:
        fn(*args, **kw)
    except exc as e:
        check(name, True, str(e)[:90])
        return e
    except Exception as e:  # noqa: BLE001
        check(name, False, f"wrong exception {type(e).__name__}: {e}")
    else:
        check(name, False, "no exception")
    return None


# ── fixtures ────────────────────────────────────────────────────────────────
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PID = os.getpid()

# Inside the AI sandbox: ~/Downloads is a default allowed root (same precedent as
# file_manager_cleanup_smoke.py). Outside it: a folder under the repo's tests/,
# which no allowed root covers — the group boundary has to open it on its own.
SANDBOX_DIR = os.path.join(os.path.expanduser("~/Downloads"), f"xlsx_actions_test_{PID}")
OUTSIDE_DIR = os.path.join(REPO, "tests", f"_xlsx_outside_{PID}")


def build_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Title"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = "Status"
    ws.append(["v1", "ok"])
    ws.append(["v2", "ok"])
    # A template row: bordered but empty. ws.max_row counts it, a human does not.
    ws["A4"].border = Border(left=Side(style="thin"))
    ws.column_dimensions["A"].width = 30
    other = wb.create_sheet("Other")
    other["A1"] = "keep"
    other["B2"] = "=1+1"
    wb.save(path)


def build_cached_formula_workbook(path):
    """build_workbook() plus the cached result Excel would have stored for
    Other!B2. openpyxl writes <v></v> for a formula, so the sheet XML is patched
    the way Excel saves it — the only way to get a cached value without Excel."""
    import io
    import zipfile

    build_workbook(path)
    src = zipfile.ZipFile(path)
    buf = io.BytesIO()
    patched = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and b"<f>1+1</f><v></v>" in data:
                data = data.replace(b"<f>1+1</f><v></v>", b"<f>1+1</f><v>2</v>")
                patched += 1
            dst.writestr(item, data)
    src.close()
    assert patched == 1, "formula cell not found in the sheet XML"
    with open(path, "wb") as f:
        f.write(buf.getvalue())


def write_csv(path, text):
    with open(path, "w", newline="", encoding="utf-8") as f:
        f.write(text)


def read_csv_lines(path):
    with open(path, "r", encoding="utf-8") as f:
        return [ln for ln in f.read().splitlines() if ln != ""]


# ── stand-in group_context (spec §3 signatures) ─────────────────────────────
def install_stub_group_context(groups_by_id):
    mod = types.ModuleType("tubecli.core.group_context")
    order = {"read": 0, "append": 1, "write": 2, "manage": 3}

    def _canon(p):
        return os.path.realpath(os.path.normpath(os.path.abspath(os.path.expanduser(str(p)))))

    def load(gid):
        return groups_by_id.get(gid)

    def effective_groups(agent_id, group_id=""):
        if group_id:
            g = groups_by_id.get(group_id)
            return [g] if g else []
        return [g for g in groups_by_id.values() if agent_id in (g.get("agents") or [])]

    def allows(have, need):
        return order.get(str(have), 0) >= order.get(str(need), 0)

    def resolve_xlsx(groups, path):
        try:
            rp = _canon(path)
        except Exception:
            return None
        for g in groups:
            for f in g.get("files") or []:
                if _canon(f["path"]) == rp:
                    return {"path": rp, "access": f.get("access") or "write"}
            for d in g.get("folders") or []:
                base = _canon(d["path"]).rstrip(os.sep)
                if rp.startswith(base + os.sep):
                    return {"path": rp, "access": d.get("access") or "write"}
        return None

    mod.ACCESS_ORDER = order
    mod.load = load
    mod.effective_groups = effective_groups
    mod.allows = allows
    mod.resolve_xlsx = resolve_xlsx
    sys.modules["tubecli.core.group_context"] = mod
    # `from tubecli.core import group_context` prefers the package attribute over
    # sys.modules once the real module has been imported, so pin both.
    import tubecli.core as core
    core.group_context = mod
    return mod


def run(coro):
    return asyncio.run(coro)


def main():
    os.makedirs(SANDBOX_DIR, exist_ok=True)
    os.makedirs(OUTSIDE_DIR, exist_ok=True)
    xlsx = os.path.join(SANDBOX_DIR, "plan.xlsx")
    csv_path = os.path.join(SANDBOX_DIR, "log.csv")
    outside_xlsx = os.path.join(OUTSIDE_DIR, "ke_hoach.xlsx")
    build_workbook(xlsx)
    build_workbook(outside_xlsx)
    write_csv(csv_path, "a,b\r\n1,2")  # no trailing newline on purpose

    svc = file_service  # the AI sandbox singleton; ~/Downloads is inside it

    # ── append keeps everything else ────────────────────────────────────────
    r = svc.append_sheet_rows(xlsx, "Data", [["v3", "new"], ["v4", 5]])
    check("append: lands after last filled row, not after the bordered blank", r["first_row"] == 4 and r["last_row"] == 5, str(r))
    wb = load_workbook(xlsx)
    ws = wb["Data"]
    check("append: values written", ws["A4"].value == "v3" and ws["B5"].value == 5)
    check("append: header bold survives", ws["A1"].font.bold is True)
    check("append: border on template row survives", ws["A4"].border.left.style == "thin")
    check("append: column width survives", ws.column_dimensions["A"].width == 30)
    check("append: other sheet + formula survive", wb["Other"]["A1"].value == "keep" and wb["Other"]["B2"].value == "=1+1")
    check("append: sheet list unchanged", wb.sheetnames == ["Data", "Other"])
    wb.close()

    r = svc.append_sheet_rows(xlsx, "data", ["v5", "x"])  # flat list = one row, case-insensitive tab
    check("append: flat list is one row, tab name case-insensitive", r["first_row"] == 6 and r["rows_added"] == 1 and r["sheet"] == "Data", str(r))

    expect_raises("append: unknown sheet lists the real ones", ValueError, svc.append_sheet_rows, xlsx, "Nope", [["x"]])
    expect_raises("append: empty rows refused", ValueError, svc.append_sheet_rows, xlsx, "Data", [])
    txt = os.path.join(SANDBOX_DIR, "plan.txt")
    write_csv(txt, "not a sheet")
    expect_raises("append: non-spreadsheet refused", ValueError, svc.append_sheet_rows, txt, None, [["x"]])

    # ── write cells ─────────────────────────────────────────────────────────
    r = svc.update_sheet_cells(xlsx, "Data", {"b2": "done", "C1": "Note"})
    check("write: cells set (lowercase ref accepted)", r["cells_written"] == 2 and r["targets"] == ["B2", "C1"], str(r))
    r = svc.update_sheet_cells(xlsx, "Data", None, rows=[[1, 2], [3, 4]], start="D1")
    check("write: block range reported", r["range"] == "D1:E2", str(r))
    wb = load_workbook(xlsx)
    ws = wb["Data"]
    check("write: values on disk", ws["B2"].value == "done" and ws["C1"].value == "Note" and ws["D1"].value == 1 and ws["E2"].value == 4)
    check("write: untouched cells intact", ws["A2"].value == "v1" and ws["A6"].value == "v5" and ws["A1"].font.bold is True)
    check("write: other sheet intact", wb["Other"]["B2"].value == "=1+1")
    wb.close()
    expect_raises("write: bad cell ref refused", ValueError, svc.update_sheet_cells, xlsx, "Data", {"1A": "x"})
    expect_raises("write: nothing to write refused", ValueError, svc.update_sheet_cells, xlsx, "Data", None)
    expect_raises("write: csv cannot be cell-edited", ValueError, svc.update_sheet_cells, csv_path, None, {"A1": "x"})

    # ── read with truncation ────────────────────────────────────────────────
    r = svc.read_sheet_rows(xlsx, "Data", max_rows=2)
    check("read: truncated to max_rows but total known", len(r["rows"]) == 2 and r["total_rows"] == 6 and r["truncated"] is True, str({k: r[k] for k in ("total_rows", "truncated")}))
    check("read: header row + sheet names", r["rows"][0][:3] == ["Title", "Status", "Note"] and r["sheets"] == ["Data", "Other"] and r["sheet"] == "Data")
    r = svc.read_sheet_rows(xlsx, None, max_rows=500)
    check("read: default sheet is the active one, no truncation", r["sheet"] == "Data" and r["truncated"] is False and len(r["rows"]) == 6)
    check("read: rows padded to equal width", len({len(x) for x in r["rows"]}) == 1)

    # ── formula cells after an in-place edit ────────────────────────────────
    # openpyxl never writes cached results, so the appends above left every
    # formula in the workbook without a value. data_only=True alone would show
    # a blank where the owner's total is; the read must show the formula text
    # and say so, and must still prefer a cached result when the file has one.
    r = svc.read_sheet_rows(xlsx, "Other")
    check("read: formula without cached result shows the formula text",
          r["rows"][1][1] == "=1+1" and r["formula_cells"] == 1 and r["formulas_uncomputed"] == 1, str(r["rows"]))
    r = svc.read_sheet_rows(xlsx, "Other", max_rows=1)
    check("read: uncomputed count covers shown rows only", r["formula_cells"] == 1 and r["formulas_uncomputed"] == 0, str(r))
    cached = os.path.join(SANDBOX_DIR, "cached.xlsx")
    build_cached_formula_workbook(cached)
    r = svc.read_sheet_rows(cached, "Other")
    check("read: cached formula result wins over the formula text",
          r["rows"][1][1] == "2" and r["formula_cells"] == 1 and r["formulas_uncomputed"] == 0, str(r["rows"]))
    svc.update_sheet_cells(cached, "Other", {"A2": "edited"})
    r = svc.read_sheet_rows(cached, "Other")
    check("read: the cached result is gone after an in-place edit (why the fallback exists)",
          r["rows"][1][1] == "=1+1" and r["formulas_uncomputed"] == 1, str(r["rows"]))
    r = svc.read_sheet_rows(xlsx, "Data")
    check("read: sheet without formulas reports none", r["formula_cells"] == 0 and r["formulas_uncomputed"] == 0)

    # ── CSV ─────────────────────────────────────────────────────────────────
    r = svc.append_sheet_rows(csv_path, None, [["3", "4"], ["5", "x,y"]])
    check("csv: append after missing trailing newline", r["first_row"] == 3 and r["last_row"] == 4, str(r))
    lines = read_csv_lines(csv_path)
    check("csv: rows on disk, quoting kept", lines == ["a,b", "1,2", "3,4", '5,"x,y"'], str(lines))
    r = svc.read_sheet_rows(csv_path, None, max_rows=10)
    check("csv: read", r["total_rows"] == 4 and r["rows"][3] == ["5", "x,y"] and r["sheet"] == "", str(r["rows"]))

    # ── sandbox still applies to the AI service ─────────────────────────────
    expect_raises("sandbox: outside path refused by file_service", ValueError, svc.read_sheet_rows, outside_xlsx)
    extra = FileService(extra_roots=[OUTSIDE_DIR])
    check("sandbox: extra root opens it", extra.read_sheet_rows(outside_xlsx)["total_rows"] == 3)

    # ── handlers ────────────────────────────────────────────────────────────
    from tubecli.extensions.file_manager.extension import FileManagerExtension

    ext = FileManagerExtension()
    actions = ext.get_telegram_actions()
    check("handlers registered", all(k in actions for k in ("xlsx_read", "xlsx_append", "xlsx_write")) and "drive_list" in actions)

    # No groups at all → sandbox behaviour.
    install_stub_group_context({})
    ctx = {"agent": {"id": "a1", "name": "Bot"}, "lang": "en"}
    out = run(actions["xlsx_read"]({"path": xlsx, "sheet": "Data", "max_rows": 3}, ctx))
    check("no-group read: table rendered", out.startswith("📊") and "| Title | Status | Note" in out and "more row(s)" in out, out[:160])
    out = run(actions["xlsx_append"]({"path": xlsx, "sheet": "Data", "rows": [["h1", "h2"]]}, ctx))
    check("no-group append: ok text", out.startswith("✅") and "rows 7–7" in out, out)
    out = run(actions["xlsx_write"]({"path": xlsx, "sheet": "Other", "cells": {"A2": "w"}}, ctx))
    check("no-group write: ok text", out.startswith("✅") and "A2" in out and "Other" in out, out)
    out = run(actions["xlsx_write"]({"path": xlsx, "sheet": "Other", "range": "B3:C3", "values": [["p", "q"]]}, ctx))
    check("no-group write: gsheet-style range/values accepted", out.startswith("✅") and "B3:C3" in out, out)
    out = run(actions["xlsx_read"]({"path": xlsx, "sheet": "Other"}, ctx))
    check("no-group read: uncomputed formula cells are called out", "| =1+1 |" in out and "formula cell(s)" in out, out[-200:])
    out = run(actions["xlsx_read"]({"path": outside_xlsx}, ctx))
    check("no-group read: outside sandbox refused", out.startswith("❌"), out[:120])
    out = run(actions["xlsx_read"]({}, ctx))
    check("missing path refused", out.startswith("❌"), out)
    out = run(actions["xlsx_read"]({"path": xlsx}, {"agent": {"id": "a1"}, "lang": "vi"}))
    check("vi text for vi bot language", "hiện" in out, out[:80])

    # Groups: the manifest is the boundary.
    install_stub_group_context({
        "g1": {
            "group_id": "g1", "label": "Content", "agents": ["a1"],
            "files": [{"alias": "Kế hoạch tuần.xlsx", "path": outside_xlsx, "ext": "xlsx", "access": "append"}],
            "folders": [], "sheets": [],
        },
        "g2": {
            "group_id": "g2", "label": "Ops", "agents": ["a2"],
            "files": [], "folders": [{"path": OUTSIDE_DIR, "access": "write"}], "sheets": [],
        },
    })
    out = run(actions["xlsx_read"]({"path": outside_xlsx, "sheet": "Data"}, ctx))
    check("group read: path outside the sandbox but in the group works", out.startswith("📊") and "| v1 | ok |" in out, out[:160])
    out = run(actions["xlsx_read"]({"path": "kế hoạch tuần.xlsx"}, ctx))
    check("group read: alias resolves to the group file", out.startswith("📊"), out[:120])
    out = run(actions["xlsx_append"]({"path": outside_xlsx, "sheet": "Data", "rows": [["g", "row"]]}, ctx))
    check("group append: access=append allows append", out.startswith("✅"), out)
    out = run(actions["xlsx_write"]({"path": outside_xlsx, "sheet": "Data", "cells": {"A1": "x"}}, ctx))
    check("group write: access=append refuses write", out.startswith("❌") and "'write'" in out and "'append'" in out, out)
    out = run(actions["xlsx_read"]({"path": xlsx}, ctx))
    check("group read: sandbox file NOT in the group is refused", out.startswith("❌") and "not shared" in out, out[:120])
    wb = load_workbook(outside_xlsx)
    check("group write refused → file untouched", wb["Data"]["A1"].value == "Title" and wb["Data"]["A4"].value == "g")
    wb.close()

    # Explicit group_ids from the pipeline win over membership.
    out = run(actions["xlsx_write"]({"path": outside_xlsx, "sheet": "Data", "cells": {"C3": "ops"}},
                                    {"agent": {"id": "stranger"}, "group_ids": ["g2"], "lang": "en"}))
    check("group_ids: folder entry with write access allows write", out.startswith("✅") and "C3" in out, out)
    out = run(actions["xlsx_read"]({"path": outside_xlsx}, {"agent": {"id": "stranger"}, "lang": "en"}))
    check("no membership, no group_ids → sandbox refuses outside path", out.startswith("❌"), out[:120])
    out = run(actions["xlsx_read"]({"path": xlsx}, {"agent": {"id": "a1"}, "group_id": "g2", "lang": "en"}))
    check("explicit group_id without the file → refused", out.startswith("❌") and "not shared" in out, out[:120])


if __name__ == "__main__":
    try:
        main()
    finally:
        shutil.rmtree(SANDBOX_DIR, ignore_errors=True)
        shutil.rmtree(OUTSIDE_DIR, ignore_errors=True)
    print(f"\n{PASS} passed, {FAIL} failed")
    sys.exit(1 if FAIL else 0)
